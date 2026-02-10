
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
import pyperclip
import qrcode
import webbrowser
import csv
import requests
import openpyxl
import shutil
import time
import math # Added for charts
import customtkinter as ctk # MODERN UI
from PIL import Image, ImageTk
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import threading
import logging
import traceback
import re

# Log yozishni sozlash
logging.basicConfig(filename='app.log', level=logging.DEBUG, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Fayl sozlamalari
# Fayl sozlamalari
DB_FILE = "mahalla_bazasi.json"
TRASH_FILE = "trash.json"
BACKUP_DIR = "backups"
LOG_FILE = "activity_log.json"
SETTINGS_FILE = "settings.json"

import time
import shutil

class DataManager:
    def __init__(self):
        self.data = self.load_json(DB_FILE)
        self.trash = self.load_json(TRASH_FILE)
        self.categories = self.load_json("categories.json")
        self.activity_log = self.load_json(LOG_FILE)
        self.settings = self.load_json(SETTINGS_FILE)
        
        # Standart Sozlamalar
        if not self.settings:
            self.settings = {"font_size": 15}
            
        if not self.categories:
             self.categories = ["Mahalla (MFY)", "Maktab", "Bog'cha (MTT)", "Boshqa"]
        self.ensure_backup_dir()

    def ensure_backup_dir(self):
        if not os.path.exists(BACKUP_DIR):
            os.makedirs(BACKUP_DIR)

    def load_json(self, filepath):
        if os.path.exists(filepath):
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    return json.load(f)
            except: return []
        return []

    def save_data(self):
        self.save_json(DB_FILE, self.data)

    def save_trash(self):
        self.save_json(TRASH_FILE, self.trash)

    def save_categories(self):
        self.save_json("categories.json", self.categories)

    def save_settings(self):
        self.save_json(SETTINGS_FILE, self.settings)

    def log_activity(self, user, action, details):
        try:
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            entry = {
                "time": timestamp,
                "user": user,
                "action": action,
                "details": details
            }
            self.activity_log.insert(0, entry)
            if len(self.activity_log) > 1000:
                self.activity_log.pop()
            self.save_json(LOG_FILE, self.activity_log)
        except: pass

    def save_json(self, filepath, data):
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

    def move_to_trash(self, item):
        if item in self.data:
            self.data.remove(item)
            item["deleted_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
            self.trash.append(item)
            self.save_data()
            self.save_trash()
            return True
        return False

    def restore_from_trash(self, item):
        if item in self.trash:
            self.trash.remove(item)
            if "deleted_at" in item: del item["deleted_at"]
            self.data.append(item)
            self.save_data()
            self.save_trash()
            return True
        return False

    def permanent_delete(self, item):
        if item in self.trash:
            self.trash.remove(item)
            self.save_trash()
            return True
        return False

    def backup_data(self):
        if os.path.exists(DB_FILE):
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(BACKUP_DIR, f"backup_{timestamp}.json")
            try:
                shutil.copy2(DB_FILE, backup_path)
                # Faqat oxirgi 10 ta nusxani saqlash
                backups = sorted([os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR)])
                while len(backups) > 10:
                    os.remove(backups.pop(0))
            except: pass

class MahallaDasturi:
    def __init__(self, root):
        self.root = root
        self.root.title("Pop Tumani Smart Boshqaruv Tizimi (PRO)")
        self.root.geometry("1350x800")
        self.root.configure(bg="#f4f7f6")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        
        self.data_manager = DataManager()
        self.data = self.data_manager.data
        self.font_size = self.data_manager.settings.get("font_size", 15)
        self.filtered_data = self.data[:]
        self.setup_ui()

    def on_close(self):
        self.data_manager.backup_data()
        self.root.destroy()

    def save_data(self):
        self.data_manager.save_data()

    def load_sync_config(self):
        default_url = "https://docs.google.com/spreadsheets/d/1l4MVpVGoyWMP_9Px9QG4V3hWLdQf9LlJ/edit"
        try:
            with open("sync_config.json", "r") as f:
                return json.load(f).get("sheet_id", default_url)
        except: return default_url

    def save_sync_config(self, sheet_id):
        with open("sync_config.json", "w") as f:
            json.dump({"sheet_id": sheet_id}, f)

    def setup_ui(self):
        # MAVZU SOZLAMALARI: Zamonaviy Palitra (Slate & Ko'k)
        self.themes = {
            "light": {
                "bg": "#f8fafc", "fg": "#334155", 
                "content_bg": "#f1f5f9", 
                "sidebar": "#1e293b", "sidebar_text": "#e2e8f0", 
                "card_bg": "white", "text": "#1e293b", "accent": "#3b82f6"
            },
            "dark": {
                "bg": "#0f172a", "fg": "#e2e8f0", 
                "content_bg": "#1e293b", 
                "sidebar": "#020617", "sidebar_text": "#94a3b8", 
                "card_bg": "#1e293b", "text": "#f8fafc", "accent": "#60a5fa"
            }
        }
        self.current_theme = "light"
        self.admin_password = "123" # Default
        self.last_auth_time = 0; self.auth_timeout = 20 * 60
        
        # Sinxronizatsiya sozlamalarini yuklash
        self.sheet_identifier = self.load_sync_config()
        
        # Modern Font
        self.lbl_font = ("Segoe UI", self.font_size)
        self.head_font = ("Segoe UI", int(self.font_size * 2.1), "bold") # Katta Sarlavha
        
        # Absolyut Yo'l (Xatolikni oldini olish uchun)
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.icon_path = os.path.join(self.base_dir, "popdata.png")
        
        # Set Icon
        try:
            icon_img = ImageTk.PhotoImage(file=self.icon_path)
            self.root.iconphoto(False, icon_img)
            self.root.title("POP Tuman") 
        except Exception as e: print(f"Icon Load Fail: {e}")
        self.btn_font = ("Segoe UI", int(self.font_size * 0.9), "bold")
        
        # Foydalanuvchi Rollari
        self.current_role = None
        self.users = {"admin": "123", "operator": "1"}

        # ASOSIY MAKET (LAYOUT)
        self.main_container = ctk.CTkFrame(self.root, corner_radius=0, fg_color=("white", "#1a1a1a"))
        self.main_container.pack(fill="both", expand=True)

        # SIDEBAR
        self.sidebar = tk.Frame(self.main_container, bg=self.themes["light"]["sidebar"], width=260)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False) # Qat'iy kenglik
        
        # LOGO AREA
        try:
             # Load Image Absolute
             pil_img = Image.open(self.icon_path)
             pil_img = pil_img.resize((150, 150)) # Resize Bigger
             logo_photo = ImageTk.PhotoImage(pil_img)
             
             self.logo_lbl = tk.Label(self.sidebar, image=logo_photo, bg=self.themes["light"]["sidebar"], pady=10)
             self.logo_lbl.image = logo_photo # Havolani saqlab qolish
             self.logo_lbl.pack(pady=(20, 10))
             
             tk.Label(self.sidebar, text="POP TUMANI\nSMART TIZIM", fg="white", bg=self.themes["light"]["sidebar"], font=("Segoe UI", 16, "bold")).pack(fill="x")
        except Exception as e:
             print(f"Logo Fail: {e}")
             self.logo_lbl = tk.Label(self.sidebar, text="POP TUMANI\nSMART TIZIM", fg="white", bg=self.themes["light"]["sidebar"], font=("Segoe UI", 16, "bold"), pady=30)
             self.logo_lbl.pack(fill="x")
        
        # NAVIGATSIYA
        ctk.CTkLabel(self.sidebar, text="ASOSIY", font=("Segoe UI", 12, "bold"), text_color="#95a5a6", anchor="w").pack(fill="x", padx=30, pady=(10,5))
        self.create_sidebar_btn("📊 Dashboard", self.show_dashboard)
        self.create_sidebar_btn("📋 Ro'yxat", self.show_table)
        self.create_sidebar_btn("⚙ Sozlamalar", self.show_settings)
        
        ctk.CTkLabel(self.sidebar, text="TIZIM", font=("Segoe UI", 12, "bold"), text_color="#95a5a6", anchor="w").pack(fill="x", padx=30, pady=(20,5))
        self.create_sidebar_btn("🗑 Chiqindi Qutisi", self.show_trash)
        self.create_sidebar_btn("☁ Cloud Sync", self.open_cloud_menu)
        
        # Pastki boshqaruv tugmalari
        self.btn_theme = self.create_sidebar_btn("🌙 Tungi Rejim", self.toggle_theme)
        ctk.CTkFrame(self.sidebar, height=2, fg_color="#34495e").pack(fill="x", padx=20, pady=10) # Ajratuvchi chiziq
        
        # Sync Status Label
        self.lbl_sync = ctk.CTkLabel(self.sidebar, text="☁ Integratsiya", text_color="gray", font=("Segoe UI", 11))
        self.lbl_sync.pack(fill="x", pady=(0, 5))

        self.create_sidebar_btn("🚪 Chiqish", self.on_close, text_color="#ef4444")

        # TARKIB QISMI (CONTENT AREA)
        self.content_area = tk.Frame(self.main_container, bg=self.themes["light"]["content_bg"])
        self.content_area.pack(side="right", fill="both", expand=True)

        # Holatni saqlash
        self.current_view = None
        self.cat_var = tk.StringVar(value="Barchasi") # Global filter variable
        self.style = ttk.Style()
        self.update_style() # Initial style
        self.show_dashboard()

    def create_sidebar_btn(self, text, cmd, fg_color="transparent", hover_color="#34495e", text_color=None):
        btn = ctk.CTkButton(self.sidebar, text=text, command=cmd, 
                            fg_color=fg_color, hover_color=hover_color, text_color=text_color if text_color else "white",
                            font=("Segoe UI", 16), anchor="w", height=45, corner_radius=8)
        btn.pack(fill="x", padx=15, pady=5)
        return btn

    def check_password(self):
        # 1. Vaqtni tekshirish (Session check)
        current_time = time.time()
        if (current_time - self.last_auth_time) < self.auth_timeout:
            return True # 20 daqiqa o'tmagan, ruxsat beriladi
        
        # 2. Modern Password Dialog (Blocking)
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Xavfsizlik")
        dialog.geometry("340x220")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center Dialog
        x = self.root.winfo_x() + (self.root.winfo_width()//2) - 170
        y = self.root.winfo_y() + (self.root.winfo_height()//2) - 110
        dialog.geometry(f"+{x}+{y}")
        
        ctk.CTkLabel(dialog, text="🔑 Tizimga kirish", font=("Segoe UI", 16, "bold")).pack(pady=(20, 10))
        ctk.CTkLabel(dialog, text="Davom etish uchun parolni kiriting", font=("Segoe UI", 12), text_color="gray").pack()
        
        entry = ctk.CTkEntry(dialog, show="*", width=220, height=35, font=("Segoe UI", 14), placeholder_text="Parol...")
        entry.pack(pady=15)
        entry.focus()
        
        self.password_result = None
        def on_confirm(event=None):
            self.password_result = entry.get()
            dialog.destroy()
            
        entry.bind("<Return>", on_confirm)
        ctk.CTkButton(dialog, text="Tasdiqlash", command=on_confirm, width=220, height=35, font=("Segoe UI", 12, "bold")).pack(pady=5)
        
        self.root.wait_window(dialog)
        
        # Verify
        role = None
        for r, p in self.users.items():
            if self.password_result == p:
                role = r
                break
        
        if role:
            self.last_auth_time = time.time()
            self.current_role = role
            self.show_toast(f"Muvaffaqiyatli kirildi! ({role.upper()})")
            return True
        else:
            if self.password_result is not None:
                messagebox.showerror("Xato", "Parol noto'g'ri!")
            return False

    def toggle_theme(self):
        # Toggle Mode
        current_mode = ctk.get_appearance_mode()
        if current_mode == "Dark":
            ctk.set_appearance_mode("Light")
            self.current_theme = "light"
            self.btn_theme.configure(text="🌙 Tungi Rejim")
        else:
            ctk.set_appearance_mode("Dark")
            self.current_theme = "dark"
            self.btn_theme.configure(text="☀ Kunduzi Rejim")
            
        # Manually update Treeview
        self.update_treeview_style()
        
        # Refresh View
        if self.current_view == "dashboard": self.show_dashboard()
        elif self.current_view == "table": self.show_table()
        elif self.current_view == "trash": self.show_trash()
        
        # Update Styles for Treeview
    def update_treeview_style(self):
        # Oddiy Treeview ranglarini CTk ga moslashtirish kerak
        mode = ctk.get_appearance_mode()
        
        # Colors for Light/Dark
        bg = "#2b2b2b" if mode == "Dark" else "white"
        fg = "white" if mode == "Dark" else "#333333" 
        field = "#2b2b2b" if mode == "Dark" else "white"
        header_bg = "#34495e" if mode == "Dark" else "#e5e7eb"
        header_fg = "white" if mode == "Dark" else "#1f2937"
        
        self.style.theme_use("clam")
        
        # Configure Rows
        self.style.configure("Treeview", 
                             background=bg, 
                             foreground=fg, 
                             fieldbackground=field, 
                             rowheight=45, # O'qish qulayligi uchun balandroq qatorlar
                             borderwidth=0, 
                             font=("Segoe UI", 14)) # Kattaroq Shrift
                             
        # Configure Header
        self.style.configure("Treeview.Heading", 
                             background=header_bg, 
                             foreground=header_fg, 
                             relief="flat",
                             font=("Segoe UI", 13, "bold"))
                             
        self.style.map("Treeview", background=[("selected", "#3b82f6")])



    def clear_content(self):
        for widget in self.content_area.winfo_children():
            widget.destroy()

    def show_dashboard(self):
        self.clear_content()
        self.current_view = "dashboard"
        
        # MAIN SCROLLABLE CONTAINER
        scroll_dash = ctk.CTkScrollableFrame(self.content_area, fg_color="transparent")
        scroll_dash.pack(fill="both", expand=True)

        t = self.themes[self.current_theme]
        
        # Header Frame
        head_frame = ctk.CTkFrame(scroll_dash, fg_color=t["content_bg"])
        head_frame.pack(fill="x", padx=40, pady=(30,20))
        
        # Text
        tk.Label(head_frame, text="Boshqaruv Paneli", font=self.head_font, bg=t["content_bg"], fg=t["text"]).pack(side="left")
        
        # Logo Image in Dashboard
        try:
            dash_img = Image.open(self.icon_path).resize((100, 100))
            dash_photo = ImageTk.PhotoImage(dash_img)
            lbl = tk.Label(head_frame, image=dash_photo, bg=t["content_bg"])
            lbl.image = dash_photo 
            lbl.pack(side="right")
        except: pass
        
        # ASOSIY AJRATILGAN OYNA (Chap: Kartalar, O'ng: Diagramma)
        split_frame = ctk.CTkFrame(scroll_dash, fg_color=t["content_bg"])
        split_frame.pack(fill="x", padx=30)
        
        # CHAP: STATISTIKA KARTALARI
        left_side = ctk.CTkFrame(split_frame, fg_color=t["content_bg"])
        left_side.pack(side="left", fill="both", expand=True)

        # 1. Jami (Katta Karta - To'liq Eniga)
        self.create_modern_card(left_side, "Jami Tashkilotlar", len(self.data), "#2c3e50", "🏢", pady=10)
        
        # Grid Container uchun Frame
        grid_frame = ctk.CTkFrame(left_side, fg_color="transparent")
        grid_frame.pack(fill="both", expand=True, pady=10)
        
        # Grid ustunlarini moslash (3 ta ustun)
        grid_frame.grid_columnconfigure(0, weight=1)
        grid_frame.grid_columnconfigure(1, weight=1)
        grid_frame.grid_columnconfigure(2, weight=1)

        # Dinamik Kategoriyalar
        colors = ["#10b981", "#f59e0b", "#8b5cf6", "#3b82f6", "#e11d48", "#14b8a6", "#f97316", "#6366f1"]
        chart_data = []

        # Mavjud kategoriyalar bo'yicha hisoblash
        total_categorized = 0
        for idx, cat in enumerate(self.data_manager.categories):
             # Filter logic mirrors filter_data STRICT match + Legacy
             count = 0
             for i in self.data:
                 s_val = str(i.get("s", "")).strip()
                 is_match = (s_val == cat)
                 # Legacy
                 if not is_match:
                    if cat == "Mahalla (MFY)" and s_val in ["Mahalla", "MFY"]: is_match = True
                    elif cat == "Maktab" and s_val in ["Maktablar"]: is_match = True
                    elif cat == "Bog'cha (MTT)" and s_val in ["MTT", "Bog'cha"]: is_match = True
                 
                 if is_match: count += 1
             
             total_categorized += count
             col = colors[idx % len(colors)]
             chart_data.append((cat, count, col))
             
             # Icon selection
             icon = "📌"
             if "Mahalla" in cat: icon = "🏘"
             elif "Maktab" in cat: icon = "🏫"
             elif "Bog'cha" in cat: icon = "🧸"
             
             # Grid Card yaratish
             r = idx // 3
             c = idx % 3
             self.create_grid_card(grid_frame, cat, count, col, icon, r, c)
             
        # Add "Boshqa" category if data is missing from chart
        total_items = len(self.data)
        if total_categorized < total_items:
            diff = total_items - total_categorized
            chart_data.append(("Boshqa (Kategoriyasiz)", diff, "#95a5a6"))

        # O'NG: DIAGRAMMA (DONUT CHART) + LEGEND
        right_side = tk.Frame(split_frame, bg=t["content_bg"])
        right_side.pack(side="right", padx=20, fill="y")
        
        self.draw_donut_chart(right_side, chart_data)

        # So'nggi Faoliyat
        tk.Label(scroll_dash, text="So'nggi Faoliyat", font=("Segoe UI", int(self.font_size * 1.2), "bold"), bg=t["content_bg"], fg=t["text"]).pack(anchor="w", padx=40, pady=(40,15))
        
        # Log konteyneri
        log_frame = tk.Frame(scroll_dash, bg=t["card_bg"], highlightbackground="#e2e8f0", highlightthickness=1)
        log_frame.pack(fill="both", expand=True, padx=40, pady=(0, 40))
        
        row_col = "#f1f5f9" if self.current_theme == "light" else "#334155"
        
        # Show actual last 5 logs from DataManager or dummy if empty
        recent_logs = self.data_manager.activity_log[:5] if self.data_manager.activity_log else []
        
        if not recent_logs:
             for i in range(3): # Dummy
                f = tk.Frame(log_frame, bg=t["card_bg"])
                f.pack(fill="x", pady=1)
                tk.Label(f, text="• Tizim ishga tushdi", font=("Segoe UI", int(self.font_size * 0.9)), bg=t["card_bg"], fg=t["text"]).pack(side="left", padx=15, pady=12)
                tk.Frame(log_frame, height=1, bg=row_col).pack(fill="x")
        else:
            for log in recent_logs:
                f = tk.Frame(log_frame, bg=t["card_bg"])
                f.pack(fill="x", pady=1)
                txt = f"• {log.get('user', '?').upper()}: {log.get('action')} - {log.get('details')}"
                tk.Label(f, text=txt, font=("Segoe UI", int(self.font_size * 0.9)), bg=t["card_bg"], fg=t["text"]).pack(side="left", padx=15, pady=12)
                tk.Label(f, text=log.get("time", "")[-8:], font=("Segoe UI", int(self.font_size * 0.7)), bg=t["card_bg"], fg="#94a3b8").pack(side="right", padx=15)
                tk.Frame(log_frame, height=1, bg=row_col).pack(fill="x")

    def draw_donut_chart(self, parent, data):
        # Canvas asosidagi zamonaviy diagramma
        t = self.themes[self.current_theme]
        sz = 340 # KATTALASHTIRILDI (220 -> 340)
        canvas = tk.Canvas(parent, width=sz, height=sz, bg=t["content_bg"], bd=0, highlightthickness=0)
        canvas.pack()
        
        total = sum(d[1] for d in data)
        if total == 0: 
            canvas.create_text(sz/2, sz/2, text="Ma'lumot yo'q", font=("Segoe UI", 12), fill="gray")
            return
        
        start_deg = 90
        center = sz/2
        radius = 120 # KATTALASHTIRILDI (80 -> 120)
        width = 40 # KATTALASHTIRILDI (25 -> 40)
        
        # LEGENDA ORQALI KO'RSATISH
        legend_frame = tk.Frame(parent, bg=t["content_bg"])
        legend_frame.pack(pady=(10, 0))
        
        for name, val, col in data:
            if val == 0: continue
            
            # Foizni hisoblash
            percent = (val / total) * 100
            extent = (val / total) * 360
            
            # Ark chizish
            tag_name = f"slice_{name}"
            safe_tag = "".join(x for x in tag_name if x.isalnum())
            
            canvas.create_arc(center-radius, center-radius, center+radius, center+radius, 
                              start=start_deg, extent=-extent, style="arc", outline=col, width=width, tags=(safe_tag, "slice"))
            
            canvas.tag_bind(safe_tag, "<Button-1>", lambda e, n=name: self.filter_from_chart(n))
            canvas.tag_bind(safe_tag, "<Enter>", lambda e, c=canvas, t=safe_tag: c.itemconfigure(t, width=width+5))
            canvas.tag_bind(safe_tag, "<Leave>", lambda e, c=canvas, t=safe_tag: c.itemconfigure(t, width=width))
            
            start_deg -= extent
            
            # Legenda qatori (Rang - Nom - Soni - Foiz)
            l_row = tk.Frame(legend_frame, bg=t["content_bg"])
            l_row.pack(anchor="w", pady=1)
            
            tk.Frame(l_row, bg=col, width=10, height=10).pack(side="left", padx=(0, 5))
            tk.Label(l_row, text=f"{name}:", font=("Segoe UI", int(self.font_size*0.9), "bold"), bg=t["content_bg"], fg=t["text"]).pack(side="left")
            tk.Label(l_row, text=f"{val} ({percent:.1f}%)", font=("Segoe UI", int(self.font_size*0.9)), bg=t["content_bg"], fg="#64748b").pack(side="left", padx=5)

        # Markaziy Matn
        canvas.create_text(center, center-15, text="Statistika", font=("Segoe UI", int(self.font_size*0.9), "bold"), fill="#94a3b8")
        canvas.create_text(center, center+20, text=f"{total}", font=("Segoe UI", int(self.font_size*1.8), "bold"), fill=t["text"])
        
        # Yordamchi matn
        tk.Label(parent, text="(Bo'limni ko'rish uchun diagrammaga bosing)", font=("Segoe UI", int(self.font_size*0.5)), bg=t["content_bg"], fg="#94a3b8").pack(pady=5)

    def filter_from_chart(self, category):
        # Interactive filter - Direct Match
        self.cat_var.set(category)
        self.show_table() # Switch to table
        self.filter_data() # Apply filter
        self.show_toast(f"{category} bo'yicha saralandi!")

    def show_toast(self, message):
        # Zamonaviy bloklanmagan bildirishnoma
        toast = tk.Toplevel(self.root)
        toast.overrideredirect(True)
        toast.geometry(f"300x50+{self.root.winfo_x() + self.root.winfo_width() - 320}+{self.root.winfo_y() + self.root.winfo_height() - 70}")
        toast.configure(bg="#333")
        
        tk.Label(toast, text=message, fg="white", bg="#333", font=("Segoe UI", int(self.font_size * 0.7))).pack(expand=True, fill="both")
        
        # Animation: Fade in/out (simulated by destroy after time)
        toast.after(3000, toast.destroy)

    def create_modern_card(self, parent, title, value, color, icon, pady=5):
        t = self.themes[self.current_theme]
        # Soya/Chegara uchun tashqi ramka
        card_border = tk.Frame(parent, bg="#e2e8f0" if self.current_theme == "light" else "#334155", padx=1, pady=1)
        card_border.pack(side="top", fill="x", expand=True, pady=pady)
        
        # Ichki karta
        card = tk.Frame(card_border, bg=t["card_bg"])
        card.pack(fill="both", expand=True)
        
        # Chap rangli chiziq
        tk.Frame(card, bg=color, width=4).pack(side="left", fill="y")
        
        # Content
        content = tk.Frame(card, bg=t["card_bg"], padx=15, pady=15)
        content.pack(fill="both", expand=True)
        
        # Belgi (Icon) aylanasi
        icon_lbl = tk.Label(content, text=icon, font=("Segoe UI", 18), bg=t["card_bg"], fg=color)
        icon_lbl.pack(side="left", anchor="center")
        
        # Text
        info = tk.Frame(content, bg=t["card_bg"])
        info.pack(side="left", padx=(15, 0))
        tk.Label(info, text=title, font=("Segoe UI", int(self.font_size * 0.6), "bold"), fg="#64748b", bg=t["card_bg"]).pack(anchor="w")
        tk.Label(info, text=str(value), font=("Segoe UI", int(self.font_size * 1.2), "bold"), fg=t["text"], bg=t["card_bg"]).pack(anchor="w")

    def create_grid_card(self, parent, title, value, color, icon, row, col):
        t = self.themes[self.current_theme]
        # Frame
        card = ctk.CTkFrame(parent, fg_color=t["card_bg"], corner_radius=8)
        
        # Click handler
        def on_click(e):
            self.filter_from_chart(title)
            
        # Left Color Bar (Fake border effect)
        bar = tk.Frame(card, bg=color, width=4)
        bar.pack(side="left", fill="y", padx=(0, 5))
        
        # Icon
        lbl_icon = tk.Label(card, text=icon, font=("Segoe UI", 18), bg=t["card_bg"], fg=color)
        lbl_icon.pack(side="left", padx=5)
        
        # Info
        info_frame = tk.Frame(card, bg=t["card_bg"])
        info_frame.pack(side="left", fill="both", expand=True, pady=10, padx=5)
        
        lbl_title = tk.Label(info_frame, text=title, font=("Segoe UI", int(self.font_size * 0.85), "bold"), fg="#64748b", bg=t["card_bg"], anchor="w")
        lbl_title.pack(fill="x")
        
        lbl_val = tk.Label(info_frame, text=str(value), font=("Segoe UI", int(self.font_size * 1.4), "bold"), fg=t["text"], bg=t["card_bg"], anchor="w")
        lbl_val.pack(fill="x")
        
        # Bind events
        for w in [card, bar, lbl_icon, info_frame, lbl_title, lbl_val]:
            w.bind("<Button-1>", on_click)
            try: w.configure(cursor="hand2")
            except: pass
            
        card.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)
        return card

    def show_trash(self):
        self.clear_content()
        self.current_view = "trash"
        
        # SARLAVHA (Mavzu qo'llab-quvvatlaydigan ranglar bilan)
        ctk.CTkLabel(self.content_area, text="Chiqindi Qutisi (Trash)", font=("Segoe UI", int(self.font_size * 1.7), "bold"), text_color=("#c0392b", "#e74c3c")).pack(anchor="w", padx=30, pady=20)
        ctk.CTkLabel(self.content_area, text="O'chirilgan ma'lumotlarni tiklashingiz yoki butunlay o'chirishingiz mumkin", font=("Segoe UI", int(self.font_size * 0.8)), text_color="gray").pack(anchor="w", padx=35)

        # JADVAL KONTEYNERI
        tree_frame = ctk.CTkFrame(self.content_area)
        tree_frame.pack(fill="both", expand=True, padx=30, pady=(10, 20))

        # JADVAL
        self.update_treeview_style()
        self.trash_tree = ttk.Treeview(tree_frame, columns=("m","f","date"), show="headings")
        
        self.trash_tree.heading("m", text="Tashkilot Nomi"); self.trash_tree.column("m", width=300)
        self.trash_tree.heading("f", text="Rahbar"); self.trash_tree.column("f", width=200)
        self.trash_tree.heading("date", text="O'chirilgan vaqt"); self.trash_tree.column("date", width=150)
        
        self.trash_tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # MA'LUMOTLARNI YUKLASH
        for i in self.data_manager.trash:
            self.trash_tree.insert("", "end", values=(i.get("m"), i.get("f"), i.get("deleted_at","-")))

        # AMALLAR
        btn_frame = ctk.CTkFrame(self.content_area, fg_color="transparent")
        btn_frame.pack(fill="x", padx=30, pady=20)
        
        ctk.CTkButton(btn_frame, text="♻ Tiklash", fg_color="#27ae60", height=40, font=("Segoe UI", 12, "bold"), command=self.restore_item).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="🔥 Butunlay O'chirish", fg_color="#c0392b", height=40, hover_color="#a93226", font=("Segoe UI", 12, "bold"), command=self.perm_delete_item).pack(side="left", padx=5)

    def restore_item(self):
        sel = self.trash_tree.selection()
        if not sel: return
        for s in sel:
            v = self.trash_tree.item(s)["values"]
            # Find in trash by name and director (simple match)
            item = next((i for i in self.data_manager.trash if i.get("m") == v[0] and i.get("f") == v[1]), None)
            if item:
                self.data_manager.restore_from_trash(item)
        self.show_trash()
        self.filter_data() # Agar ko'rinib turgan bo'lsa, asosiy jadvalni yangilash
        messagebox.showinfo("OK", "Ma'lumotlar tiklandi!")

    def perm_delete_item(self):
        sel = self.trash_tree.selection()
        if not sel: return
        if not messagebox.askyesno("Diqqat", "Rostdan ham butunlay o'chirmoqchimisiz? Qaytarib bo'lmaydi!"): return
        
        for s in sel:
            v = self.trash_tree.item(s)["values"]
            item = next((i for i in self.data_manager.trash if i.get("m") == v[0] and i.get("f") == v[1]), None)
            if item:
                self.data_manager.permanent_delete(item)
        self.show_trash()

        self.show_trash()

    def show_settings(self):
        # XAVFSIZLIK TEKSHIRUVI
        if not self.check_password(): return

        # QO'SHIMCHA HUBQUQ TEKSHIRUVI (Faqat Admin)
        if self.current_role != "admin":
            # Custom Admin Password Dialog
            dialog = ctk.CTkToplevel(self.root)
            dialog.title("Admin Tasdiqlash")
            dialog.geometry("350x200")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.resizable(False, False)
            
            # Center
            x = self.root.winfo_x() + (self.root.winfo_width()//2) - 175
            y = self.root.winfo_y() + (self.root.winfo_height()//2) - 100
            dialog.geometry(f"+{x}+{y}")
            
            ctk.CTkLabel(dialog, text="🔒 Admin Ruxsati", font=("Segoe UI", 16, "bold"), text_color="#e74c3c").pack(pady=(20, 10))
            ctk.CTkLabel(dialog, text="Sozlamalarga kirish uchun Admin parolini kiriting:", font=("Segoe UI", 12)).pack()
            
            entry = ctk.CTkEntry(dialog, show="*", width=220, height=35)
            entry.pack(pady=10)
            entry.focus()
            
            self.admin_pwd_res = None
            def on_confirm(e=None):
                self.admin_pwd_res = entry.get()
                dialog.destroy()
                
            entry.bind("<Return>", on_confirm)
            ctk.CTkButton(dialog, text="Kirish", command=on_confirm, width=220, fg_color="#e74c3c", hover_color="#c0392b").pack(pady=5)
            
            self.root.wait_window(dialog)
            
            if self.admin_pwd_res == self.users["admin"]:
                self.current_role = "admin"
                self.show_toast("Admin rejimi faollashdi!")
            else:
                if self.admin_pwd_res is not None: messagebox.showerror("Xato", "Parol noto'g'ri!")
                return

        self.clear_content()
        self.current_view = "settings"
        
        ctk.CTkLabel(self.content_area, text="Sozlamalar", font=("Segoe UI", 26, "bold"), text_color=("#2c3e50", "white")).pack(anchor="w", padx=30, pady=20)
        
        # Tabs
        tabs = ctk.CTkTabview(self.content_area)
        tabs.pack(fill="both", expand=True, padx=30, pady=(0, 20))
        
        tab_cat = tabs.add("Kategoriyalar")
        tab_font = tabs.add("Ko'rinish (Shrift)")
        tab_log = tabs.add("Tarix (Logs)")
        tab_tools = tabs.add("Asboblar (Tools)")
        
        # --- TAB 1: Categories ---
        # Add Frame
        add_frame = ctk.CTkFrame(tab_cat)
        add_frame.pack(fill="x", padx=10, pady=10)
        
        self.new_cat_var = tk.StringVar()
        ctk.CTkLabel(add_frame, text="Yangi tur qo'shish:", font=("Segoe UI", 12, "bold")).pack(side="left", padx=10, pady=10)
        ctk.CTkEntry(add_frame, textvariable=self.new_cat_var, width=200, height=35, placeholder_text="Masalan: Hokim Yordamchilari").pack(side="left", padx=5)
        
        def add_cat():
            new_c = self.new_cat_var.get().strip()
            if not new_c: return
            if new_c in self.data_manager.categories:
                self.show_toast("Bu tur allaqachon mavjud!")
                return
            
            self.data_manager.categories.append(new_c)
            self.data_manager.save_categories()
            self.data_manager.log_activity(self.current_role, "Kategoriya Qo'shildi", f"Nomi: {new_c}")
            self.new_cat_var.set("")
            self.show_toast(f"'{new_c}' qo'shildi!")
            refresh_list()
        
        ctk.CTkButton(add_frame, text="+ Qo'shish", command=add_cat, fg_color="#27ae60", width=100, height=35).pack(side="left", padx=10)

        # List
        list_frame = ctk.CTkFrame(tab_cat)
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        scroll = ctk.CTkScrollableFrame(list_frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=5, pady=5)
        
        def refresh_list():
            for w in scroll.winfo_children(): w.destroy()
            for cat in self.data_manager.categories:
                r = ctk.CTkFrame(scroll, fg_color=("white", "#333"), height=40)
                r.pack(fill="x", pady=2)
                
                ctk.CTkLabel(r, text=cat, font=("Segoe UI", 12)).pack(side="left", padx=15, pady=5)
                
                if cat not in ["Mahalla (MFY)", "Maktab", "Bog'cha (MTT)"]: 
                    ctk.CTkButton(r, text="🗑", width=30, height=30, fg_color="#c0392b", command=lambda c=cat: delete_cat(c)).pack(side="right", padx=5, pady=5)
        
        def delete_cat(cat):
            if messagebox.askyesno("O'chirish", f"'{cat}' turini o'chirmoqchimisiz?"):
                self.data_manager.categories.remove(cat)
                self.data_manager.save_categories()
                self.data_manager.log_activity(self.current_role, "Kategoriya O'chirildi", f"Nomi: {cat}")
                refresh_list()
                self.show_toast("O'chirildi!")
        
        refresh_list()
        
        # --- TAB 2: Font Size ---
        font_frame = ctk.CTkFrame(tab_font, fg_color="transparent")
        font_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(font_frame, text="Dastur Shrifti O'lchami:", font=("Segoe UI", 18, "bold")).pack(pady=10)
        ctk.CTkLabel(font_frame, text="O'zgartirishlar faqat dasturni qayta ishga tushirgandan so'ng to'liq qo'llaniladi.", font=("Segoe UI", 12), text_color="gray").pack(pady=(0, 20))
        
        def set_font(size):
            self.data_manager.log_activity(self.current_role, "Shrif O'zgardi", f"Yangi o'lcham: {size}")
            # Update Font and UI
            self.change_font_size(size)
            self.show_toast(f"Shrift {size} ga o'zgardi!")
            
        btn_box = ctk.CTkFrame(font_frame, fg_color="transparent")
        btn_box.pack(pady=10)
        
        sizes = [("Kichik", 12), ("O'rta (Standard)", 15), ("Katta", 18), ("Juda Katta", 22)]
        for label, sz in sizes:
            col = "#3b82f6" if self.font_size == sz else "gray"
            ctk.CTkButton(btn_box, text=f"{label} ({sz})", command=lambda s=sz: set_font(s), width=150, height=45, fg_color=col, font=("Segoe UI", 14)).pack(side="left", padx=10)

        # Manual Input
        manual_frame = ctk.CTkFrame(font_frame, fg_color="transparent")
        manual_frame.pack(pady=20)
        
        ctk.CTkLabel(manual_frame, text="Yoki aniq o'lchamni kiriting:", font=("Segoe UI", 14)).pack(side="left", padx=10)
        
        manual_var = tk.StringVar(value=str(self.font_size))
        entry_font = ctk.CTkEntry(manual_frame, textvariable=manual_var, width=80, font=("Segoe UI", 14))
        entry_font.pack(side="left", padx=5)
        
        def save_manual():
            try:
                val = int(manual_var.get())
                if val < 8 or val > 50:
                    self.show_toast("O'lcham 8 va 50 orasida bo'lishi kerak!")
                    return
                set_font(val)
            except:
                self.show_toast("Iltimos, raqam kiriting!")
        
        ctk.CTkButton(manual_frame, text="✅ Qo'llash", command=save_manual, width=100, fg_color="#27ae60").pack(side="left", padx=10)

        # --- TAB 3: Activity Log ---
        log_frame = ctk.CTkFrame(tab_log)
        log_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        cols = ("time", "user", "action", "details")
        log_tree = ttk.Treeview(log_frame, columns=cols, show="headings", height=20)
        
        log_tree.heading("time", text="Vaqt"); log_tree.column("time", width=150, anchor="center")
        log_tree.heading("user", text="Foydalanuvchi"); log_tree.column("user", width=120, anchor="center")
        log_tree.heading("action", text="Harakat"); log_tree.column("action", width=180, anchor="w")
        log_tree.heading("details", text="Tafsilotlar"); log_tree.column("details", width=400, anchor="w")
        
        log_tree.pack(fill="both", expand=True)
        
        # Scrollbar
        vsb = ttk.Scrollbar(log_frame, orient="vertical", command=log_tree.yview)
        vsb.place(relx=1, rely=0, relheight=1, anchor="ne")
        log_tree.configure(yscrollcommand=vsb.set)
        
        # Load Logs
        logs = self.data_manager.activity_log
        for log in logs:
            log_tree.insert("", "end", values=(log.get("time"), log.get("user"), log.get("action"), log.get("details")))

        # --- TAB 4: Tools (Asboblar) ---
        tools_frame = ctk.CTkFrame(tab_tools, fg_color="transparent")
        tools_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(tools_frame, text="Ma'lumotlar Bazasi Asboblari", font=("Segoe UI", 18, "bold")).pack(pady=10)
        
        # 1. FIND DUPLICATES
        dup_frame = ctk.CTkFrame(tools_frame)
        dup_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(dup_frame, text="🔍 Dublikatlarni Topish (Bir xil INN)", font=("Segoe UI", 14, "bold")).pack(side="left", padx=15, pady=15)
        
        def find_duplicates():
            # Scan
            inn_map = {}
            for item in self.data:
                inn = item.get("inn")
                if inn and inn.isdigit(): # Valid INN only
                    if inn not in inn_map: inn_map[inn] = []
                    inn_map[inn].append(item)
            
            duplicates = {k:v for k,v in inn_map.items() if len(v) > 1}
            
            if not duplicates:
                self.show_toast("Dublikatlar topilmadi! Baza toza.")
                return
                
            # Show Window
            dw = ctk.CTkToplevel(self.root)
            dw.title(f"Topildi: {len(duplicates)} ta guruh")
            dw.geometry("600x500")
            
            ctk.CTkLabel(dw, text=f"⚠️ {len(duplicates)} ta INN bo'yicha takrorlanishlar topildi", font=("Segoe UI", 16, "bold"), text_color="#e74c3c").pack(pady=10)
            
            scroll = ctk.CTkScrollableFrame(dw)
            scroll.pack(fill="both", expand=True, padx=10, pady=10)
            
            for inn, items in duplicates.items():
                g_frame = ctk.CTkFrame(scroll, fg_color="transparent") # Group
                g_frame.pack(fill="x", pady=5)
                
                ctk.CTkLabel(g_frame, text=f"INN: {inn} ({len(items)} ta)", font=("Segoe UI", 13, "bold")).pack(anchor="w")
                
                for it in items:
                    r = ctk.CTkFrame(g_frame, border_width=1, border_color="gray")
                    r.pack(fill="x", padx=10, pady=2)
                    
                    info = f"{it.get('m')} | {it.get('f')}"
                    ctk.CTkLabel(r, text=info, font=("Segoe UI", 12)).pack(side="left", padx=5)
                    
                    # Delete Button per Item
                    def delete_dup(target=it, w=r):
                         if messagebox.askyesno("O'chirish", f"Chindan ham '{target.get('m')}' ni o'chirmoqchimisiz?"):
                             self.data_manager.move_to_trash(target)
                             self.data.remove(target) # Xotira ro'yxatidan to'g'ridan-to'g'ri o'chirish
                             w.destroy()
                             self.data_manager.log_activity(self.current_role, "Dublikat O'chirildi", f"{target.get('m')}")
                             self.save_data()
                    
                    ctk.CTkButton(r, text="🗑", width=30, height=30, fg_color="#c0392b", command=delete_dup).pack(side="right", padx=5, pady=2)

        ctk.CTkButton(dup_frame, text="Tekshirish", command=find_duplicates, font=("Segoe UI", 13), width=150).pack(side="right", padx=15)

    def show_table(self):
        self.clear_content()
        self.current_view = "table"
        
        # Mavzuga qarab ranglarni aniqlash
        bg = self.themes[self.current_theme]["content_bg"]
        
        # ASBOBLAR PANELI (TOOLBAR)
        ctrl = ctk.CTkFrame(self.content_area, fg_color="transparent")
        ctrl.pack(fill="x", padx=20, pady=(20, 10))
        
        # Qidiruv Konteyneri
        search_frame = ctk.CTkFrame(ctrl, fg_color="transparent")
        search_frame.pack(side="left", fill="x")

        # Filtrlash Turi
        self.f_type = ctk.CTkComboBox(search_frame, values=["Nomi", "F.I.SH", "INN", "Izoh"], width=120, height=40, font=("Segoe UI", 12))
        self.f_type.set("Nomi")
        self.f_type.pack(side="left", padx=(0, 10))
        # CTk ComboBox command logic if needed (or just query it)

        # Qidiruv maydoni
        self.s_var = tk.StringVar(); self.s_var.trace("w", self.filter_data)
        
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.s_var, width=300, height=40, font=("Segoe UI", 14), placeholder_text="Qidiruv...")
        search_entry.pack(side="left", padx=5)
        
        # Tozalash tugmasi
        ctk.CTkButton(search_frame, text="✖", width=40, height=40, fg_color="#e74c3c", command=lambda: self.s_var.set("")).pack(side="left", padx=5)
        
        # JAMI SONI YORLIQ (Bu yerdan olib tashlandi)
        # self.lbl_count = ctk.CTkLabel(search_frame, text="Jami: 0", font=("Segoe UI", 12, "bold"))
        # self.lbl_count.pack(side="left", padx=20)

        # Harakat tugmalari konteyneri
        btn_frame = ctk.CTkFrame(ctrl, fg_color="transparent")
        btn_frame.pack(side="right")

        # Tugma yordamchisi
        def add_btn(txt, cmd, col, icon=None):
            ctk.CTkButton(btn_frame, text=txt, command=cmd, fg_color=col, height=40, font=("Segoe UI", int(self.font_size * 0.8), "bold"), width=100).pack(side="right", padx=3)

        add_btn("📝 Izoh", self.manual_edit_comment, "#8e44ad")
        add_btn("📱 QR", self.show_qr, "#e67e22")
        add_btn("✈ Telegram", self.send_telegram, "#0088cc")
        add_btn("📊 Excel", self.open_export_menu, "#107c41")
        add_btn("✏ Tahrir", self.edit_item, "#f39c12")
        
        # Qo'shish tugmasi (Asosiy)
        ctk.CTkButton(btn_frame, text="+ Qo'shish", command=self.add_item, fg_color="#27ae60", height=40, font=("Segoe UI", int(self.font_size * 0.9), "bold"), width=120).pack(side="right", padx=10)

        # KATEGORIYA TABLARI (Horizontal Scroll)
        cat_scroll = ctk.CTkScrollableFrame(self.content_area, orientation="horizontal", height=50, fg_color="transparent")
        cat_scroll.pack(fill="x", padx=20, pady=10)
        
        if not hasattr(self, 'cat_var'): self.cat_var = tk.StringVar(value="Barchasi")
        
        self.cat_buttons = {}
        tabs = ["Barchasi"] + self.data_manager.categories
        
        def update_tab_ui():
            cur = self.cat_var.get()
            for t_name, btn in self.cat_buttons.items():
                if t_name == cur:
                    btn.configure(fg_color=("#3b82f6", "#2563eb"), text_color="white")
                else:
                    btn.configure(fg_color=("#e2e8f0", "#334155"), text_color=("black", "white"))

        def on_tab_click(val):
            self.cat_var.set(val)
            self.filter_data()
            update_tab_ui()

        for cat in tabs:
            w = max(80, len(cat)*10 + 20)
            btn = ctk.CTkButton(cat_scroll, text=cat, width=w, height=35,
                                font=("Segoe UI", 12, "bold"),
                                command=lambda c=cat: on_tab_click(c))
            btn.pack(side="left", padx=5)
            self.cat_buttons[cat] = btn
            
        update_tab_ui() # Apply initial style

        # JADVAL KONTEYNERI
        tree_frame = ctk.CTkFrame(self.content_area)
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # TABLE
        self.update_treeview_style()        # Stil yangilanganligiga ishonch hosil qilish
        
        # Note: Scrollbar must be packed BEFORE treeview if using side="right" on both, or AFTER if using fill.
        # Standard: Scrollbar right fill Y, Tree left fill both expand.
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        vsb.pack(side="right", fill="y")
        
        self.tree = ttk.Treeview(tree_frame, columns=("num", "s","m","f","t","inn","izoh"), show="headings", yscrollcommand=vsb.set)
        
        headers = [("num", "№", 35), ("s","Turi", 80), ("m","Tashkilot Nomi", 300), ("f","F.I.SH", 200), 
                   ("t","Tel", 100), ("inn","INN", 100), ("izoh","Izoh (Enter=Tahrir)", 200)]
        
        for col, name, width in headers:
            self.tree.heading(col, text=name, command=lambda c=col: self.sort_treeview(c, False))
            self.tree.column(col, width=width, anchor="center" if col != "m" else "w")

        self.tree.pack(fill="both", expand=True) # Scrollbar is already packed to right
        
        vsb.configure(command=self.tree.yview)

        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", self.on_double_click) 
        self.tree.bind("<Return>", self.edit_comment_inline) # Enter key binding

        # PASTKI RAMKA (Jami hisobi)
        footer = ctk.CTkFrame(self.content_area, fg_color="transparent", height=30)
        footer.pack(fill="x", padx=30, pady=(0, 10))
        self.lbl_count = ctk.CTkLabel(footer, text="Jami: 0", font=("Segoe UI", int(self.font_size * 0.9), "bold"), text_color="gray")
        self.lbl_count.pack(side="right")

        self.update_table(self.filtered_data) # Hozirgi filtrlangan ma'lumotlarni yuklash

    def filter_data_seg(self, value):
        self.cat_var.set(value)
        self.filter_data()
        
    def manual_edit_comment(self):
        # Satr ichida tahrirlash uchun tugma toetkisi
        self.edit_comment_inline(None)

    def edit_comment_inline(self, event):
        # Get selected item
        sel = self.tree.focus()
        if not sel: return
        
        # Ensure item is visible
        self.tree.see(sel)
        self.root.update_idletasks() # Maket hisoblanishini ta'minlash
        
        # "izoh" ustuni katakchasi koordinatalarini olish
        try:
            bbox = self.tree.bbox(sel, "izoh")
            if not bbox: return # Ustun ko'rinmayapti
            x, y, w, h = bbox
        except: return 
        
        # Kirish vidjeti qatlami (overlay) yaratish
        # Frame helps with border visibility if needed, but direct Entry is usually enough
        entry = tk.Entry(self.tree, font=("Segoe UI", self.font_size))
        entry.place(x=x, y=y, width=w, height=h)
        
        # Hozirgi qiymatni o'rnatish
        current_val = self.tree.item(sel)["values"][6] # Izoh is @ index 6 (0=num, 1=s, 2=m, 3=f, 4=t, 5=inn, 6=izoh)
        entry.insert(0, str(current_val))
        entry.select_range(0, tk.END)
        entry.focus_force() # Fokusni majburlash

        def save_edit(event):
            new_text = entry.get()
            # 1. UI ni yangilash
            current_values = list(self.tree.item(sel)["values"])
            current_values[6] = new_text
            self.tree.item(sel, values=current_values)
            
            # 2. Ma'lumot manbasini yangilash
            inn = str(current_values[5]) # INN is @ index 5
            item = next((i for i in self.data if str(i.get("inn")) == inn), None)
            if item:
                item["izoh"] = new_text
                self.data_manager.save_data()
            
            entry.destroy()
            try:
                self.tree.focus_set() # Fokusni daraxtga qaytarish
                
                # Keyingi elementga o'tish (Excel uslubi)
                next_item = self.tree.next(sel)
                if next_item:
                    self.tree.selection_set(next_item)
                    self.tree.focus(next_item)
                    self.tree.see(next_item)
                else:
                    self.tree.focus(sel) # Stay if last
            except: pass
            
            self.sync_background() # Avto Sinxronlash

        def cancel_edit(event):
            entry.destroy()
            self.tree.focus_set() # Fokusni daraxtga qaytarish

        entry.bind("<Return>", save_edit) # Enter saves
        entry.bind("<Escape>", cancel_edit) # Esc cancels

    def open_cloud_menu(self):
        # Yon panel tugmasi uchun mini menyu
        win = tk.Toplevel(self.root)
        win.title("Cloud"); win.geometry("300x150")
        win.configure(bg=self.themes[self.current_theme]["content_bg"]) # Theme fix
        
        tk.Label(win, text="GitHub Cloud Sinxronlash", font=("Arial", 12, "bold"), bg=self.themes[self.current_theme]["content_bg"], fg=self.themes[self.current_theme]["text"]).pack(pady=10)
        tk.Button(win, text="☁ Ma'lumotni Yuklash (Upload)", bg="#27ae60", fg="white", command=self.upload_to_github).pack(fill="x", padx=20, pady=5)
        tk.Button(win, text="☁ Ma'lumotni Olish (Download)", bg="#e67e22", fg="white", command=self.download_from_github).pack(fill="x", padx=20, pady=5)

    def open_export_menu(self):
         # Mini menu implementation
        win = tk.Toplevel(self.root)
        win.title("Export"); win.geometry("300x200")
        win.configure(bg=self.themes[self.current_theme]["content_bg"]) # Theme fix

        tk.Button(win, text="📥 Joriy Jadvalni Excelga Yuklash", bg="#1f618d", fg="white", command=self.export_excel_pro).pack(fill="x", padx=20, pady=5)
        tk.Button(win, text="📊 Google Sheet", bg="#107c41", fg="white", command=self.export_to_gsheet_dummy).pack(fill="x", padx=20, pady=5)
        tk.Button(win, text="🔄 Google Sheet Sync (Real)", bg="#2ecc71", fg="white", command=self.open_gsheet_sync_menu).pack(fill="x", padx=20, pady=5)

    def open_gsheet_sync_menu(self):
        if self.current_role != "admin": # Cheklov
            messagebox.showerror("Ruxsat Yo'q", "Faqat ADMIN!"); return

        win = tk.Toplevel(self.root)
        win.title("Sozlamalar")
        win.geometry("350x250")
        
        ctk.CTkLabel(win, text="Google Sheets Integratsiyasi", font=("Segoe UI", 16, "bold")).pack(pady=10)
        ddddddddddddddddddddddd                                                          
        # Kalit fayl holati
        key_status = "✅ Fayl mavjud" if os.path.exists("service_account.json") else "❌ Fayl yo'q"
        lbl_status = ctk.CTkLabel(win, text=f"Key File: {key_status}", text_color="green" if "mavjud" in key_status else "red")
        lbl_status.pack()
        
        def select_key():
            path = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
            if path:
                try:
                    shutil.copy(path, "service_account.json")
                    lbl_status.configure(text="✅ Fayl mavjud", text_color="green")
                    messagebox.showinfo("OK", "Kalit fayl o'rnatildi!")
                except Exception as e: messagebox.showerror("Xato", str(e))
        ddddddd
        if not os.path.exists("service_account.json"):
            ctk.CTkButton(win, text="🔑 Kalit Faylni Tanlash", command=select_key).pack(pady=10)
        else:
            ctk.CTkLabel(win, text="Avtomatik Sinxronizatsiya: YOQILGAN 🟢", font=("Segoe UI", 14, "bold"), text_color="#2ecc71").pack(pady=20)
            ctk.CTkLabel(win, text="Barcha o'zgarishlar o'zi saqlanadi.", font=("Segoe UI", 12), text_color="gray").pack()
            
            # Small re-auth button
            ctk.CTkButton(win, text="🔑 Kalitni Yangilash", command=select_key, height=24, width=120, fg_color="#7f8c8d").pack(pady=(20, 0))

        # --- NEW: Sheet ID/Link Config ---
        ctk.CTkLabel(win, text="Google Sheet Linki (ID):", font=("Segoe UI", 12, "bold")).pack(pady=(15, 5))
        
        self.entry_sheet_id = ctk.CTkEntry(win, width=300)
        self.entry_sheet_id.pack(pady=5)
        self.entry_sheet_id.insert(0, self.sheet_identifier)
        
        def save_sheet_id():
            new_val = self.entry_sheet_id.get().strip()
            if not new_val: return
            self.sheet_identifier = new_val
            self.save_sync_config(new_val)
            messagebox.showinfo("Saqlandi", "Yangi Sheet Linki saqlandi!")
            
        ctk.CTkButton(win, text="Floppy Disk 💾 Saqlash", command=save_sheet_id, fg_color="#27ae60").pack(pady=10)


    def sync_background(self):

        # Avto Sinxronlash ishga tushiruvchisi
        threading.Thread(target=self.do_sync, args=("upload", self.sheet_identifier, True), daemon=True).start()

    def do_sync(self, direction, sheet_name="MahallaBazasi", silent=False):
        if not os.path.exists("service_account.json"):
            if not silent: messagebox.showerror("Xato", "Kalit fayl topilmadi!")
            return
        
        # Agar qo'lda sinxronlash bo'lsa, konfigni saqlash
        if not silent:
            self.sheet_identifier = sheet_name
            self.save_sync_config(sheet_name)
        
        if silent:
            self.lbl_sync.configure(text="☁ Yuklanmoqda...", text_color="#f39c12")
        
        try:
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            creds = ServiceAccountCredentials.from_json_keyfile_name("service_account.json", scope)
            client = gspread.authorize(creds)
            
            try: 
                 val = sheet_name.strip()
                 logging.info(f"Connecting to sheet input: {val}")
                 
                 # Regex yordamida ID ajratib olish (Eng yaxshi usul)
                 # Matches .../d/THIS_PART/...
                 match = re.search(r"/d/([a-zA-Z0-9-_]+)", val)
                 
                 if match:
                     sheet_id = match.group(1)
                     logging.info(f"Extracted Sheet ID: {sheet_id}")
                     spreadsheet = client.open_by_key(sheet_id)
                 elif len(val) > 20 and " " not in val and not val.startswith("http"):
                      # Raw ID provided
                      logging.info(f"Using raw Sheet ID: {val}")
                      spreadsheet = client.open_by_key(val)
                 elif val.startswith("http"): 
                      # Fallback for weird URLs
                      spreadsheet = client.open_by_url(val)
                 else: 
                      spreadsheet = client.open(val)
            except Exception as e: 
                logging.error(f"Sheet Open Failed: {e}")
                if direction == "upload" and not val.startswith("http"): # Only create if name provided
                    spreadsheet = client.create(val); client.insert_permission(spreadsheet.id, None, perm_type='anyone', role='reader')
                else: 
                     if silent: self.lbl_sync.configure(text="❌ Sheet topilmadi", text_color="red"); return
                     raise Exception("Sheet topilmadi (Nomini to'g'ri yozing yoki Linkni qo'ying)")

            sheet = spreadsheet.sheet1
            
            if direction == "upload":
                # Upload Logic
                data_to_upload = [["Turi", "Nomi", "Rahbar", "Tel", "INN", "Izoh", "ID"]] # Header + ID
                for i in self.data:
                        data_to_upload.append([
                            i.get("s"), i.get("m"), i.get("f"), i.get("t"), str(i.get("inn")), i.get("izoh", ""), i.get("uuid", "")
                        ])
                
                sheet.clear()
                sheet.update(data_to_upload)
                logging.info("Upload Successful")
                
                if silent:
                     self.lbl_sync.configure(text="✅ Bulutda", text_color="#2ecc71")
                else:
                    messagebox.showinfo("OK", "Ma'lumotlar Google Sheetga yuklandi!")
                    webbrowser.open(f"https://docs.google.com/spreadsheets/d/{spreadsheet.id}")
                
            elif direction == "download":
                # Download Logic
                raw_data = sheet.get_all_records()
                logging.info(f"Downloaded {len(raw_data)} records")
                new_db = []
                for row in raw_data:
                    new_db.append({
                        "s": row.get("Turi"), "m": row.get("Nomi"), "f": row.get("Rahbar"),
                        "t": row.get("Tel"), "inn": str(row.get("INN")), "izoh": row.get("Izoh")
                    })
                
                self.data = new_db
                self.data_manager.data = self.data
                self.save_data()
                self.filter_data()
                if not silent: messagebox.showinfo("OK", "Ma'lumotlar Google Sheetdan yuklab olindi!")

        except Exception as e: 
            err_msg = str(e)
            logging.error(f"Sync Error: {err_msg}\n{traceback.format_exc()}")
            
            if "operation is not supported" in err_msg or "400" in err_msg:
                 friendly_msg = "XATO: Siz Excel (.xlsx) fayl ulagansiz!\nIltimos, Google Sheetni ochib 'File -> Save as Google Sheets' qiling va yangi linkni ishlating."
                 if not silent: messagebox.showerror("Format Xatosi", friendly_msg)
                 else: print(friendly_msg) # console fallback
            elif silent:
                self.lbl_sync.configure(text="❌ Xato", text_color="red")
            else:
                messagebox.showerror("Xato", f"Google Sheet Xatosi:\n{err_msg}")
    
    def on_double_click(self, event):
        self.edit_item()

    def sort_treeview(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        try:
            # Try numeric sort if possible
            l.sort(key=lambda t: int(t[0]), reverse=reverse)
        except ValueError:
            l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)

        self.tree.heading(col, command=lambda: self.sort_treeview(col, not reverse))


    def filter_data(self, *args):
        if self.current_view != "table": return
        q = self.s_var.get().lower().strip() # Added strip()
        cat = self.cat_var.get()
        tp = self.f_type.get()
        
        res = []
        res = []
        for i in self.data:
            n = str(i.get("m", "")).lower()
            # Dynamic Filter Logic (Strict + Legacy Support)
            s_val = str(i.get("s", "")).strip()
            
            # 1. Exact Match (Primary for New Categories)
            is_match = (s_val == cat)
            
            # 2. Legacy Support for Renamed Core Categories
            if not is_match:
                if cat == "Mahalla (MFY)" and s_val in ["Mahalla", "MFY"]: is_match = True
                elif cat == "Maktab" and s_val in ["Maktablar"]: is_match = True
                elif cat == "Bog'cha (MTT)" and s_val in ["MTT", "Bog'cha"]: is_match = True
            
            match_cat = (cat == "Barchasi") or is_match
             
            if match_cat:
                target = ""
                if tp == "Nomi": target = n
                elif tp == "F.I.SH": target = str(i.get("f","")).lower()
                elif tp == "INN": target = str(i.get("inn",""))
                elif tp == "Izoh": target = str(i.get("izoh","")).lower()
                
                # Check for substring match
                if q in target: res.append(i)
                
        self.filtered_data = res
        self.update_table(res)

    def update_table(self, d_list):
        for r in self.tree.get_children(): self.tree.delete(r)
        
        # Respecting current sort would be ideal, but simply appending matches user expectations for "Filtered View"
        for idx, i in enumerate(d_list, 1):
            self.tree.insert("", "end", values=(
                str(idx), # Number
                i.get("s","-"), i.get("m","-"), i.get("f","-"), 
                i.get("t","-"), i.get("inn","-"), i.get("izoh", "")
            ))
        
        # Update Counter
        if hasattr(self, "lbl_count"):
             self.lbl_count.configure(text=f"Jami: {len(d_list)} ta")
        
        # Set focus to top item for keyboard nav
        if self.tree.get_children():
            first = self.tree.get_children()[0]
            self.tree.selection_set(first)
            self.tree.focus(first)

    # --- PRO FUNKSIYALAR ---

    def export_excel_pro(self):
        # Exports ONLY the currently filtered data (Smart Export)
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not path: return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = f"Tashkilotlar - {self.cat_var.get()}" if hasattr(self, 'cat_var') else "Tashkilotlar"
            ws.append(["Turi", "Nomi", "Rahbar", "Tel", "INN", "Izoh"])
            
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")

            # Uses self.filtered_data which contains exactly what is shown on screen (filtered by search OR category)
            for i in self.filtered_data:
                ws.append([i.get("s"), i.get("m"), i.get("f"), i.get("t"), i.get("inn"), i.get("izoh", "")])
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = max_length + 2

            wb.save(path); messagebox.showinfo("OK", f"Excel fayl saqlandi! ({len(self.filtered_data)} ta tashkilot)")
        except Exception as e: messagebox.showerror("Xato", str(e))

    def upload_to_github(self, auto=False):
        if auto:
            # Auto mode: Sillent upload
            gist_id = ""
            if os.path.exists("gist_config.txt"):
                 with open("gist_config.txt", "r") as f: gist_id = f.read().strip()
            if gist_id:
                try:
                    # Token should be saved securely, but for now we assume it's prompted or stored elsewhere?
                    # The original code only prompted for token. 
                    # If auto is True, we can't prompt. 
                    # For now, let's just skip if no token mechanism exists for auto.
                    pass 
                except: pass
            return

        token = simpledialog.askstring("GitHub Token", "GitHub Tokeningizni kiriting:")
        if not token: return
        
        gist_content = {
            "description": "Pop Tumani Bazasi", "public": False,
            "files": {"mahalla_bazasi.json": {"content": json.dumps(self.data, indent=4, ensure_ascii=False)}}
        }
        try:
            r = requests.post("https://api.github.com/gists", json=gist_content, headers={"Authorization": f"token {token}"})
            if r.status_code == 201:
                with open("gist_config.txt", "w") as f: f.write(r.json()['id'])
                messagebox.showinfo("OK", "Baza Cloudga yuklandi!")
            else: messagebox.showerror("Xato", "Internet yoki Token xatosi!")
        except Exception as e: messagebox.showerror("Xato", str(e))

    def download_from_github(self):
        gist_id = ""
        if os.path.exists("gist_config.txt"):
            with open("gist_config.txt", "r") as f: gist_id = f.read().strip()
        
        if not gist_id:
            gist_id = simpledialog.askstring("Gist ID", "Baza ID raqamini kiriting:")
        if not gist_id: return

        try:
            r = requests.get(f"https://api.github.com/gists/{gist_id}")
            if r.status_code == 200:
                self.data = json.loads(r.json()['files']['mahalla_bazasi.json']['content'])
                self.data_manager.data = self.data # Update manager data
                self.save_data(); self.filter_data()
                messagebox.showinfo("OK", "Baza yangilandi!")
            else: messagebox.showerror("Xato", "Baza topilmadi!")
        except Exception as e: messagebox.showerror("Xato", str(e))

    def export_to_gsheet_dummy(self):
        # Configdan o'qish
        val = self.load_sync_config()
        
        # Link yoki ID ekanligini tekshirish
        if "http" in val:
            url = val
        else:
            # Agar faqat ID bo'lsa, to'liq link yasaymiz
            url = f"https://docs.google.com/spreadsheets/d/{val}"
            
        # Havolani brauzerda ochish
        webbrowser.open(url)
        # messagebox.showinfo("Google Sheet", "Jadval brauzerda ochildi!")

    # --- YORDAMCHI FUNKSIYALAR ---

    def edit_item(self):
        if not self.check_password(): return # Password Protected
        sel = self.tree.focus()
        if sel:
            v = self.tree.item(sel)["values"]
            item = next((i for i in self.data if i.get("inn") == str(v[5])), None) # INN is at index 5 now
            if item: self.open_win("Tahrirlash", item)

    def add_item(self): 
        if not self.check_password(): return # Password Protected
        self.open_win("Yangi qo'shish")

    def open_win(self, title, item=None):
        win = ctk.CTkToplevel(self.root)
        win.title(title)
        win.geometry("450x600")
        win.transient(self.root) 
        win.grab_set()
        
        # Center
        x = self.root.winfo_x() + (self.root.winfo_width()//2) - 225
        y = self.root.winfo_y() + (self.root.winfo_height()//2) - 300
        win.geometry(f"+{x}+{y}")
        
        # Header
        ctk.CTkLabel(win, text=title, font=("Segoe UI", 22, "bold"), text_color=("#2c3e50", "#ecf0f1")).pack(pady=(25, 10))
        ctk.CTkLabel(win, text="ma'lumotlarni to'ldiring", font=("Segoe UI", 12), text_color="gray").pack(pady=(0, 20))

        # Fields Config
        # Dynamic Options from DataManager
        cat_opts = self.data_manager.categories
        
        form_config = [
            ("s", "Tashkilot Turi", "combo", cat_opts),
            ("m", "Tashkilot Nomi", "entry", None),
            ("f", "Rahbar (F.I.SH)", "entry", None),
            ("t", "Telefon Raqam", "entry", None),
            ("inn", "INN (Soliq to'lovchi)", "entry", None),
            ("izoh", "Qo'shimcha Izoh", "entry", None),
        ]
        
        widgets = {}
        
        widgets = {}
        
        # Scrollable Container for Fields
        container = ctk.CTkScrollableFrame(win, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        
        for key, lbl, w_type, opts in form_config:
            ctk.CTkLabel(container, text=lbl, font=("Segoe UI", 12, "bold"), anchor="w").pack(fill="x", pady=(5,0))
            
            if w_type == "combo":
                w = ctk.CTkComboBox(container, values=opts, height=35, font=("Segoe UI", 13))
            else:
                w = ctk.CTkEntry(container, height=35, font=("Segoe UI", 13))
                
            w.pack(fill="x", pady=(0, 8))
            widgets[key] = w
            
            # Pre-fill
            if item:
                val = item.get(key, "")
                if w_type == "combo": w.set(val)
                else: w.insert(0, val)
        
        # Bind Input Mask for Phone
        if "t" in widgets:
             # CTkEntry contains a specialized entry, need to bind to internal entry or wrapper. 
             # CTkEntry events are usually passed through.
             widgets["t"].bind("<KeyRelease>", self.format_phone_input)
        
        def save():
            # VALIDATION
            val_inn = widgets["inn"].get()
            val_name = widgets["m"].get()
            
            if not val_name:
                self.show_toast("Xatolik: Tashkilot nomi kiritilmadi!"); return
            if val_inn and not val_inn.isdigit(): 
                 self.show_toast("Xatolik: INN raqam bo'lishi kerak!"); return

            # DUPLICATE CHECK
            if val_inn:
                # Find any other item with same INN
                # If editing (item is not None), ignore self
                # item.get('inn') is the old INN
                exists = next((x for x in self.data if x.get("inn") == val_inn and x != item), None)
                if exists:
                    msg = f"DIQQAT: Bu INN ({val_inn}) allaqachon mavjud!\n\nTashkilot: {exists.get('m')}\nRahbar: {exists.get('f')}\n\nBaribir saqlansinmi?"
                    if not messagebox.askyesno("Dublikat Topildi", msg):
                        return

            # 1. Collect Data
            d = {}
            for key, _, w_type, _ in form_config:
                d[key] = widgets[key].get()
            
            # 2. Save
            if item: 
                self.data[self.data.index(item)] = d
                self.data_manager.log_activity(self.current_role, "Tahrirlash", f"{d.get('m')} yangilandi")
            else: 
                self.data.append(d)
                self.data_manager.log_activity(self.current_role, "Qo'shish", f"{d.get('m')} yangi qo'shildi")
            
            # 3. Persist
            self.save_data() 
            self.filter_data()
            
            # 5. Close
            win.destroy()
            self.show_toast("Muvaffaqiyatli saqlandi!")
            
            self.sync_background() # Auto Sync
        
        ctk.CTkButton(win, text="SAQLASH", command=save, height=50, font=("Segoe UI", 14, "bold"), fg_color="#27ae60", hover_color="#2ecc71").pack(fill="x", padx=40, pady=20)

    def format_phone_input(self, event):
        # Avto-format: +998 (99) 123-45-67
        entry = event.widget
        if event.keysym == "BackSpace": return # O'chirishga ruxsat
        
        text = entry.get()
        digits = "".join(filter(str.isdigit, text))
        
        # 998 ni olib tashlash (agar bo'lsa)
        if digits.startswith("998"):
            digits = digits[3:]
            
        if len(digits) > 9: digits = digits[:9] # Maksimum 9 ta raqam
        
        formatted = "+998 "
        if len(digits) > 0:
            formatted += f"({digits[:2]}"
        if len(digits) > 2:
            formatted += f") {digits[2:5]}"
        if len(digits) > 5:
            formatted += f"-{digits[5:7]}"
        if len(digits) > 7:
            formatted += f"-{digits[7:9]}"
            
        entry.delete(0, tk.END)
        entry.insert(0, formatted)

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item: 
            self.tree.selection_set(item)
            # Recreate menu to include Delete
            self.context_menu = tk.Menu(self.root, tearoff=0)
            self.context_menu.add_command(label="📞 Tel nusxalash", command=lambda: self.copy_cell(4)) # Adjusted index
            self.context_menu.add_command(label="🆔 INN nusxalash", command=lambda: self.copy_cell(5)) # Adjusted index
            self.context_menu.add_command(label="📝 Izohni nusxalash", command=lambda: self.copy_cell(6)) # Adjusted index
            self.context_menu.add_separator()
            self.context_menu.add_command(label="📋 Qatorni nusxalash", command=self.copy_row)
            self.context_menu.add_separator()
            self.context_menu.add_command(label="🧹 Belgilangan Izohlarni Tozalash", command=self.clear_comments)
            self.context_menu.add_command(label="🗑 BARCHA Izohlarni Tozalash", command=self.clear_all_comments)
            self.context_menu.add_separator()
            self.context_menu.add_command(label="🗑 Chiqindiga tashlash", command=self.delete_selected_item, foreground="red")
            
            self.context_menu.post(event.x_root, event.y_root)

    def clear_comments(self):
        # Bulk clear comments
        if self.current_role != "admin" and self.current_role != "operator":
             return # Should not happen usually

        sel = self.tree.selection()
        if not sel: return
        
        count = len(sel)
        if not messagebox.askyesno("Tasdiqlash", f"{count} ta tashkilotni izohini o'chirib tashlamoqchimisiz?"): return
        
        updated = False
        for s in sel:
             v = self.tree.item(s)["values"]
             # Find item by INN (unique)
             inn_val = str(v[5]) # INN is at index 5 now
             item = next((i for i in self.data if i.get("inn") == inn_val), None)
             if item:
                 item["izoh"] = ""
                 updated = True
        
        if updated:
            self.data_manager.save_data()
            self.filter_data()
            self.sync_background() # Auto Sync
            messagebox.showinfo("Bajarildi", "Izohlar tozalandi!")

    def clear_all_comments(self):
        if not self.check_password(): return # Password Protected
        if self.current_role != "admin": 
             messagebox.showerror("Ruxsat Yo'q", "Faqat ADMIN!"); return
             
        if not messagebox.askyesno("DIQQAT!", "Rostdan ham BARCHA tashkilotlarning izohlarini o'chirib tashlamoqchimisiz?\n\nBu amalni qaytarib bo'lmaydi!"): return
        
        for item in self.data:
            item["izoh"] = ""
            
        self.data_manager.save_data()
        self.filter_data()
        self.sync_background()
        messagebox.showinfo("Bajarildi", "Tizimdagi barcha izohlar tozalandi.")

    def delete_selected_item(self):
        if not self.check_password(): return # Password Protected
        if self.current_role != "admin":
             messagebox.showerror("Ruxsat Yo'q", "Faqat ADMIN o'chira oladi!"); return
        
        sel = self.tree.selection()
        if not sel: return
        if not messagebox.askyesno("O'chirish", "Haqiqatan ham bu ma'lumotni o'chirmoqchimisiz? (Keyinroq Trashdan tiklashingiz mumkin)"): return
        
        v = self.tree.item(sel)["values"]
        # Find item securely
        item = next((i for i in self.data if i.get("inn") == str(v[5])), None) # INN is 5
        if item:
            self.data_manager.move_to_trash(item)
            self.data_manager.log_activity(self.current_role, "Chiqindiga tashlandi", f"Nomi: {item.get('m')}")
            self.filter_data()
            self.sync_background() # Auto Sync
            messagebox.showinfo("O'chirildi", "Ma'lumot Chiqindi qutisiga joylandi.")

    def restore_item(self):
        if not self.check_password(): return # Password Protected
        if self.current_role != "admin":
             messagebox.showerror("Ruxsat Yo'q", "Faqat ADMIN tiklay oladi!"); return

        sel = self.trash_tree.selection()
        if not sel: return
        for s in sel:
            v = self.trash_tree.item(s)["values"]
            # Find in trash by name and director (simple match)
            item = next((i for i in self.data_manager.trash if i.get("m") == v[0] and i.get("f") == v[1]), None)
            if item:
                self.data_manager.restore_from_trash(item)
                self.data_manager.log_activity(self.current_role, "Tiklandi", f"Nomi: {item.get('m')}")
        self.show_trash()
        self.filter_data() # Update main table if it's visible
        self.sync_background() # Auto Sync
        messagebox.showinfo("OK", "Ma'lumotlar tiklandi!")

    def perm_delete_item(self):
        if not self.check_password(): return # Password Protected
        if self.current_role != "admin":
             messagebox.showerror("Ruxsat Yo'q", "Faqat ADMIN o'chira oladi!"); return

        sel = self.trash_tree.selection()
        if not sel: return
        if not messagebox.askyesno("Diqqat", "Rostdan ham butunlay o'chirmoqchimisiz? Qaytarib bo'lmaydi!"): return
        
        for s in sel:
            v = self.trash_tree.item(s)["values"]
            item = next((i for i in self.data_manager.trash if i.get("m") == v[0] and i.get("f") == v[1]), None)
            if item:
                self.data_manager.permanent_delete(item)
                self.data_manager.log_activity(self.current_role, "Butunlay O'chirildi", f"Nomi: {item.get('m')}")
        self.show_trash()
        self.sync_background() # Auto Sync

    def copy_cell(self, idx):
        try:
            sel = self.tree.item(self.tree.selection()[0])["values"]
            pyperclip.copy(str(sel[idx])); messagebox.showinfo("OK", "Nusxalandi!")
        except: pass

    def copy_row(self):
        try:
            sel = self.tree.item(self.tree.selection()[0])["values"]
            pyperclip.copy("\t".join(map(str, sel))); messagebox.showinfo("OK", "Nusxalandi!")
        except: pass # Xatolik bo'lsa indamaydi
    
    def send_telegram(self):
        try:
            v = self.tree.item(self.tree.focus())["values"]
            # Indeks: 0=No, 1=Turi, 2=Nomi, 3=Rahbar, 4=Tel, 5=INN, 6=Izoh
            izoh = f"\n📝 {v[6]}" if v[6] else ""
            # Format: Turi (v1), Nomi (v2), Rahbar (v3), Tel (v4), INN (v5)
            pyperclip.copy(f"🏢 {v[1]} {v[2]}\n👤 {v[3]}\n📞 {v[4]}\n🆔 {v[5]}{izoh}")
            messagebox.showinfo("OK", "Telegramga tayyor!")
        except: pass

    def show_qr(self):
        try:
            sel = self.tree.focus()
            if not sel: return
            v = self.tree.item(sel)["values"]
            
            # Telefon raqamni olish (Index 4)
            # Indeks: 0=No, 1=Turi, 2=Nomi, 3=Rahbar, 4=Tel, 5=INN, 6=Izoh
            raw_tel = str(v[4]) 
            
            # Faqat raqamlarni ajratib olish
            digits = "".join(filter(str.isdigit, raw_tel))
            
            tel = ""
            
            # MANTIQNI KUCHAYTIRISH:
            # 1. Agar raqamlar ichida "998" ketma-ketligi bo'lsa, o'sha joydan boshlab 12 ta raqamni olamiz.
            if "998" in digits:
                start_index = digits.find("998")
                # 998 dan boshlab 12 ta raqam bormi?
                if len(digits[start_index:]) >= 12:
                    tel = "+" + digits[start_index : start_index+12]
                else:
                    # 998 bor, lekin 12 ta emas (kaltaroq). Boricha olamiz.
                    tel = "+" + digits[start_index:]
            
            # 2. Agar 998 yo'q bo'lsa, lekin raqam uzunligi 9 ta bo'lsa (lokal)
            elif len(digits) == 9:
                tel = "+998" + digits
                
            # 3. Agar 998 yo'q, lekin 9 tadan ko'p bo'lsa (masalan 835... va oxirida 901234567)
            # Biz oxirgi 9 ta raqamni olamiz, chunki O'zbekiston raqami 9 ta (kodsiz)
            elif len(digits) > 9:
                tel = "+998" + digits[-9:]
                
            else:
                # Juda qisqa yoki tushunarsiz format. Boricha qoldiramiz.
                tel = digits

            # QR ma'lumotlarini tayyorlash
            # 835 muammosi: Ba'zi skanerlar "tel:" prefiksini "8335" (T-E-L) deb terib yuboradi.
            # Shuning uchun, biz faqat TOZA raqamni QR kodga joylaymiz.
            # Zamonaviy telefon kameralari o'zi avtomatik raqam ekanligini tushunadi.
            data = tel
            
            qr = qrcode.make(data)
            qr.save("t.png")
            
            # Oynani ko'rsatish
            w = tk.Toplevel(self.root)
            w.title("QR Kod (Telefon)")
            w.geometry("300x450")
            
            # Rasmni yuklash
            img = Image.open("t.png")
            img = img.resize((250, 250)) 
            i = ImageTk.PhotoImage(img)
            
            l = tk.Label(w, image=i)
            l.image = i 
            l.pack(pady=20)
            
            tk.Label(w, text=f"{v[3]}", font=("Segoe UI", 12, "bold")).pack() # Nomi
            tk.Label(w, text=f"{tel}", font=("Segoe UI", 16, "bold"), fg="#27ae60").pack(pady=5)
            tk.Label(w, text="(Skaner qiling va qo'ng'iroq tugmasini bosing)", font=("Segoe UI", 9), fg="gray").pack(pady=5)
            
            # Tushuntirish
            tk.Label(w, text="Eslatma: Agar '835' chiqsa, 'tel:' so'zini teryapti.\nHozir faqat raqamning o'zi chiqadi.", font=("Segoe UI", 8), fg="red").pack(pady=5)
            
        except Exception as e:
            messagebox.showerror("QR Xatosi", f"QR Kod yaratishda xatolik:\n{str(e)}")

    def change_font_size(self, new_size):
        self.font_size = new_size
        self.data_manager.settings["font_size"] = new_size
        self.data_manager.save_settings()
        
        # Update Derived Fonts
        self.lbl_font = ("Segoe UI", self.font_size)
        self.head_font = ("Segoe UI", int(self.font_size * 2.1), "bold")
        self.btn_font = ("Segoe UI", int(self.font_size * 0.9), "bold")
        
        # Update Treeview Style
        self.update_style()
        
        # Refresh Logic
        # 1. Refresh Sidebar Buttons (Destroy and Recreate Sidebar is hard, so we just update if possible or ignore)
        # For simplicity, we expect Sidebar to be static-ish, but let's try to update sidebar labels if we tracked them.
        # Ideally, we should restart the app or redraw everything. 
        # Easier approach: Just re-render the content area which is most important.
        
        if self.current_view == "dashboard": self.show_dashboard()
        elif self.current_view == "table": self.show_table()
        elif self.current_view == "settings": self.show_settings()
        elif self.current_view == "trash": self.show_trash()
    def update_style(self):
        self.style.configure("Treeview", font=("Segoe UI", self.font_size), rowheight=int(self.font_size*2.5))
        
        if self.current_theme == "dark":
            self.style.configure("Treeview", background="#34495e", foreground="white", fieldbackground="#34495e")
            self.style.configure("Treeview.Heading", font=("Segoe UI", self.font_size, "bold"), background="#2c3e50", foreground="white") # Dark Header
            self.style.map("Treeview", background=[("selected", "#3498db")])
        else:
            self.style.configure("Treeview", background="white", foreground="black", fieldbackground="white")
            self.style.configure("Treeview.Heading", font=("Segoe UI", self.font_size, "bold"), background="#ecf0f1", foreground="black") # Light Header
            self.style.map("Treeview", background=[("selected", "#3498db")])

if __name__ == "__main__":
    root = ctk.CTk()
    app = MahallaDasturi(root)
    root.mainloop()